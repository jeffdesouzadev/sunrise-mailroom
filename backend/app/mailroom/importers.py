from datetime import datetime
from openpyxl import load_workbook
import re
from zoneinfo import ZoneInfo


FORMAT_SUNRISE_EXPORT = "sunrise_export"
FORMAT_LEGACY = "legacy"
FORMAT_UNKNOWN = "unknown"


def normalize_header(value):
    if value is None:
        return ""

    return str(value).strip().lower()

def extract_timezone_from_header(value):
    """
    Read a timezone from a Sunrise export header such as:

        Timestamp (America/Chicago)

    Returns the IANA timezone name, or None if the header
    does not contain a valid timezone.
    """
    if not value:
        return None

    match = re.match(
        r"^timestamp\s*\(([^)]+)\)$",
        str(value).strip(),
        re.IGNORECASE,
    )

    if not match:
        return None

    timezone_name = match.group(1).strip()

    try:
        ZoneInfo(timezone_name)
    except Exception:
        return None

    return timezone_name

def detect_workbook_format(workbook):
    """
    Detect the workbook format automatically.

    Sunrise export format:
        Date of Birth | Name | Timestamp
    """

    for worksheet in workbook.worksheets:
        if worksheet.max_row < 1:
            continue

        headers = [
            normalize_header(cell.value)
            for cell in worksheet[1]
        ]

        if (
            len(headers) >= 3
            and headers[0] == "date of birth"
            and headers[1] == "name"
            and headers[2].startswith("timestamp")
        ):
            return FORMAT_SUNRISE_EXPORT

    #
    # We'll add the legacy-format detector here
    # once we inspect one of those files.
    #
    # if looks_like_legacy_format(workbook):
    #     return FORMAT_LEGACY

    return FORMAT_UNKNOWN


def parse_excel_date(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    #
    # openpyxl will normally convert real Excel dates
    # into datetime/date objects automatically, but this
    # provides some tolerance for text cells.
    #
    if isinstance(value, str):
        value = value.strip()

        for fmt in (
            "%m/%d/%Y",
            "%Y-%m-%d",
        ):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                pass

    return None


def parse_excel_timestamp(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    if isinstance(value, str):
        value = value.strip()

        formats = (
            "%m/%d/%Y %I:%M %p",
            "%m/%d/%Y %I:%M:%S %p",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
        )

        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                pass

    return None


def parse_sunrise_export(workbook):
    records = []

    for worksheet in workbook.worksheets:
        timezone_name = None

        if worksheet.max_row >= 1:
            headers = [
                cell.value
                for cell in worksheet[1]
            ]

            if len(headers) >= 3:
                timezone_name = extract_timezone_from_header(
                    headers[2]
                )

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if not any(row):
                continue

            dob_value = row[0] if len(row) > 0 else None
            name_value = row[1] if len(row) > 1 else None
            timestamp_value = row[2] if len(row) > 2 else None

            full_name = (
                str(name_value).strip()
                if name_value is not None
                else ""
            )

            date_of_birth = parse_excel_date(dob_value)
            visited_at = parse_excel_timestamp(timestamp_value)

            if not full_name or not date_of_birth or not visited_at:
                records.append({
                    "valid": False,
                    "sheet": worksheet.title,
                    "row": row_number,
                    "reason": (
                        "Missing or invalid name, date of birth, "
                        "or timestamp."
                    ),
                })
                continue

            records.append({
                "valid": True,
                "sheet": worksheet.title,
                "row": row_number,
                "full_name": full_name,
                "date_of_birth": date_of_birth,
                "visited_at": visited_at,
                "timezone_name": timezone_name,
            })

    return records

def parse_workbook(uploaded_file):
    workbook = load_workbook(
        uploaded_file,
        data_only=True,
    )

    workbook_format = detect_workbook_format(workbook)

    if workbook_format == FORMAT_SUNRISE_EXPORT:
        return {
            "format": workbook_format,
            "records": parse_sunrise_export(workbook),
        }

    if workbook_format == FORMAT_LEGACY:
        raise NotImplementedError(
            "Legacy import format is not implemented yet."
        )

    raise ValueError(
        "The workbook format could not be recognized."
    )