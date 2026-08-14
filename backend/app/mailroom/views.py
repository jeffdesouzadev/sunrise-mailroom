from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from datetime import date, datetime, time, timedelta
from io import BytesIO
from django.db import transaction
# from .system_timezone import get_system_timezone
from .system_timezone import get_mailroom_timezone
from django.utils import timezone
from .importers import parse_workbook

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

from .models import Client, Visit
from .serializers import (
    ClientSerializer,
    ClientDetailSerializer,
    VisitSerializer,
)


def home(request):
    return JsonResponse({
        "message": "Sunrise Mailroom backend is running."
    })


def health_check(request):
    return JsonResponse({
        "status": "ok",
        "service": "sunrise-mailroom-backend",
    })


@api_view(["GET", "POST"])
def client_list(request):
    if request.method == "GET":
        clients = Client.objects.all()

        dob = request.GET.get("dob")
        name = request.GET.get("name", "").strip()

        # Birthday remains the primary narrowing mechanism.
        if dob:
            clients = clients.filter(
                date_of_birth=dob
            )

        # Treat each entered name fragment as a required token.
        #
        # Example:
        #
        #   ?name=jean damme
        #
        # matches:
        #
        #   Jean Claude Van Damme
        #
        # because full_name must contain both "jean" AND "damme".
        if name:
            for token in name.split():
                clients = clients.filter(
                    full_name__icontains=token
                )

        clients = clients.order_by("full_name")

        serializer = ClientSerializer(
            clients,
            many=True,
        )

        return Response(serializer.data)

    serializer = ClientSerializer(
        data=request.data
    )

    if serializer.is_valid():
        serializer.save()

        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED,
        )

    return Response(
        serializer.errors,
        status=status.HTTP_400_BAD_REQUEST,
    )


@api_view(["GET", "PATCH", "DELETE"])
def client_detail(request, pk):
    client = get_object_or_404(
        Client,
        pk=pk,
    )

    if request.method == "GET":
        serializer = ClientDetailSerializer(client)
        return Response(serializer.data)

    if request.method == "PATCH":
        serializer = ClientSerializer(
            client,
            data=request.data,
            partial=True,
        )

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)

        return Response(
            serializer.errors,
            status=status.HTTP_400_BAD_REQUEST,
        )

    client.delete()

    return Response(
        status=status.HTTP_204_NO_CONTENT,
    )


@api_view(["POST"])
def client_visit(request, pk):
    client = get_object_or_404(
        Client,
        pk=pk,
    )

    visit = Visit.objects.create(
        client=client,
    )

    serializer = VisitSerializer(visit)

    return Response(
        serializer.data,
        status=status.HTTP_201_CREATED,
    )

@api_view(["GET"])
def export_visits(request):    
    local_timezone, timezone_label = get_mailroom_timezone()

    year_param = request.GET.get("year")
    start_param = request.GET.get("start")
    end_param = request.GET.get("end")

    #
    # Determine requested date range.
    #
    if year_param:
        try:
            year = int(year_param)
        except ValueError:
            return Response(
                {"error": "Invalid year."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)

    elif start_param and end_param:
        try:
            start_date = date.fromisoformat(start_param)
            end_date = date.fromisoformat(end_param)
        except ValueError:
            return Response(
                {"error": "Dates must use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if end_date < start_date:
            return Response(
                {"error": "End date must not be before start date."},
                status=status.HTTP_400_BAD_REQUEST,
            )

    else:
        #
        # Default to the current calendar year.
        #
        today = datetime.now(local_timezone).date()

        start_date = date(today.year, 1, 1)
        end_date = date(today.year, 12, 31)

    #
    # Convert the requested local dates into timezone-aware
    # datetimes for querying visited_at.
    #

    
    start_datetime = timezone.make_aware(
        datetime.combine(start_date, time.min),
        local_timezone,
    )

    end_datetime = timezone.make_aware(
        datetime.combine(
            end_date + timedelta(days=1),
            time.min,
        ),
        local_timezone,
    )

    visits = (
        Visit.objects
        .filter(
            visited_at__gte=start_datetime,
            visited_at__lt=end_datetime,
        )
        .select_related("client")
        .order_by("visited_at")
    )

    workbook = Workbook()

    #
    # Remove openpyxl's automatically-created sheet.
    #
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    #
    # Determine every calendar month represented by the
    # requested date range.
    #
    month_cursor = date(
        start_date.year,
        start_date.month,
        1,
    )

    month_keys = []

    while month_cursor <= end_date:
        month_keys.append(
            (month_cursor.year, month_cursor.month)
        )

        if month_cursor.month == 12:
            month_cursor = date(
                month_cursor.year + 1,
                1,
                1,
            )
        else:
            month_cursor = date(
                month_cursor.year,
                month_cursor.month + 1,
                1,
            )

    #
    # Create the sheets first.
    #
    sheets = {}

    multiple_years = start_date.year != end_date.year

    for sheet_year, sheet_month in month_keys:
        month_name = date(
            sheet_year,
            sheet_month,
            1,
        ).strftime("%B")

        if multiple_years:
            sheet_name = f"{month_name} {sheet_year}"
        else:
            sheet_name = month_name

        worksheet = workbook.create_sheet(
            title=sheet_name
        )

        # timezone_label = str(current_timezone)
        worksheet.append([
            "Date of Birth",
            "Name",
            f"Timestamp ({timezone_label})",
        ])

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 32
        worksheet.column_dimensions["C"].width = 24

        sheets[(sheet_year, sheet_month)] = worksheet
 
    #
    # Add visits to their corresponding month.
    #
    for visit in visits:
        local_visit = visit.visited_at.astimezone(
            local_timezone
        )

        worksheet = sheets.get(
            (
                local_visit.year,
                local_visit.month,
            )
        )

        if not worksheet:
            continue

        worksheet.append([
            visit.client.date_of_birth,
            visit.client.full_name,
            local_visit.replace(
                tzinfo=None,
                microsecond=0,
            ),
        ])

        # local_visit = timezone.localtime(
        #     visit.visited_at
        # )

        # worksheet = sheets.get(
        #     (
        #         local_visit.year,
        #         local_visit.month,
        #     )
        # )

        # if not worksheet:
        #     continue

        # worksheet.append([
        #     visit.client.date_of_birth,
        #     visit.client.full_name,
        #     local_visit.replace(
        #         tzinfo=None,
        #         microsecond=0,
        #     ),
        # ])

        row_number = worksheet.max_row

        worksheet.cell(
            row=row_number,
            column=1,
        ).number_format = "mm/dd/yyyy"

        worksheet.cell(
            row=row_number,
            column=3,
        ).number_format = "mm/dd/yyyy h:mm AM/PM"

    #
    # Write workbook to memory.
    #
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    if year_param:
        filename = f"sunrise-visits-{year}.xlsx"
    else:
        filename = (
            f"sunrise-visits-"
            f"{start_date.isoformat()}-"
            f"{end_date.isoformat()}.xlsx"
        )

    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = (
        f'attachment; filename="{filename}"'
    )

    return response

@api_view(["POST"])
def import_visits(request):
    uploaded_file = request.FILES.get("file")

    if not uploaded_file:
        return Response(
            {
                "error": "No Excel file was provided."
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    if not uploaded_file.name.lower().endswith(".xlsx"):
        return Response(
            {
                "error": "Only .xlsx files are supported."
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        parsed = parse_workbook(uploaded_file)

    except ValueError as exc:
        return Response(
            {
                "error": str(exc),
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    except NotImplementedError as exc:
        return Response(
            {
                "error": str(exc),
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    valid_records = [
        record
        for record in parsed["records"]
        if record["valid"]
    ]

    invalid_records = [
        record
        for record in parsed["records"]
        if not record["valid"]
    ]

    clients_created = 0
    clients_existing = 0
    visits_created = 0
    duplicates_skipped = 0

    
    local_timezone, _ = get_mailroom_timezone()

    with transaction.atomic():
        for record in valid_records:
            client, created = Client.objects.get_or_create(
                full_name=record["full_name"],
                date_of_birth=record["date_of_birth"],
            )
            if created:
                clients_created += 1
            else:
                clients_existing += 1

            
            visited_at = record["visited_at"]

            if timezone.is_naive(visited_at):
                visited_at = timezone.make_aware(
                    visited_at,
                    local_timezone,
                )

            visit_second_start = visited_at.replace(
                microsecond=0
            )

            visit_second_end = visit_second_start + timedelta(
                seconds=1
            )

            existing_visit = Visit.objects.filter(
                client=client,
                visited_at__gte=visit_second_start,
                visited_at__lt=visit_second_end,
            ).exists()

            if existing_visit:
                duplicates_skipped += 1
                continue

            Visit.objects.create(
                client=client,
                visited_at=visited_at,
            )

            visits_created += 1

    return Response({
        "format": parsed["format"],
        "rows_read": len(parsed["records"]),
        "valid_rows": len(valid_records),
        "invalid_rows": len(invalid_records),
        "clients_created": clients_created,
        "clients_existing": clients_existing,
        "visits_created": visits_created,
        "duplicates_skipped": duplicates_skipped,
        "errors": invalid_records[:50],
    })